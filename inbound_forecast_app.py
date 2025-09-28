import math
from datetime import datetime, timedelta
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

# =============================
# Tô màu web
# =============================
st.markdown(
    """
    <style>
    .stApp {
        background-color: #ffe6f0;
        font-family: "Segoe UI";
        color: #000;
    }
    h1 { color:#000; text-align:center; font-size:42px !important; }
    h2,h3,label,.stMarkdown { color:#000 !important; }

    div.stButton > button,
    .stDownloadButton button,
    [data-testid="stFileUploader"] button {
        background-color:#ff66b2; color:#000;
        border-radius:12px; border:2px solid #ff66b2;
        height:3em; font-size:18px;
        transition:0.3s;
    }
    div.stButton > button:hover,
    .stDownloadButton button:hover,
    [data-testid="stFileUploader"] button:hover {
        background:#ff3385; border-color:#ff3385;color:#000000; transform: scale(1.05);
    }

    .stAlert {
        background:#ff2f8f !important;
        color:#000 !important;
        border-left:5px solid #ff2f8f !important;
        border-radius:10px !important;
        padding:10px !important;
    }

    [data-testid="stMarkdownContainer"] code {
        background:#ffd6e8 !important;
        color:#000 !important;
        border:1px solid #ff5aa7 !important;
        border-radius:6px !important;
        padding:0 6px !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.set_page_config(page_title="Replenishment HND2025", layout="wide")
st.title("🪄 Hieu Ngan's Planner 💖")

# =============================
# Mô tả các yêu cầu đầu vào
REQUIRED_COLS = [
    "CAT", "SKU_code", "Available stock", "Upcoming stock", "Upcoming date",
    "Forecast OB/day", "Leadtime (day)", "DOC"
]

def validate_input(df: pd.DataFrame):
    df.columns = df.columns.str.strip()
    return [c for c in REQUIRED_COLS if c not in df.columns]

def safe_num(x):
    try:
        return float(x) if not pd.isna(x) else 0.0
    except Exception:
        return 0.0

# =============================
# =============================
def process_input(df_input, days_ahead=30, today=None):
    df = df_input.copy()
    df.columns = df.columns.str.strip()
    df["Upcoming date"] = pd.to_datetime(df["Upcoming date"], errors="coerce").dt.date

    today = today or pd.to_datetime("today").normalize().date()
    dates = [today + timedelta(days=i) for i in range(days_ahead + 1)]
    date_cols = [d.strftime("%Y-%m-%d") for d in dates]

    out_current, out_ordered = [], []

    for _, row in df.iterrows():
        # --- Inputs ---
        cat, sku = row.get("CAT"), row.get("SKU_code")
        available = safe_num(row.get("Available stock", 0))
        upcoming_stock = safe_num(row.get("Upcoming stock", 0))
        upcoming_date = row.get("Upcoming date")
        forecast = safe_num(row.get("Forecast OB/day", 0))
        leadtime = int(safe_num(row.get("Leadtime (day)", 0)))
        doc = safe_num(row.get("DOC", 0))

        # --- Simulation: Output.current ---
        cur, stocks_current, oos_date = available, [], None
        for d in dates:
            cur -= forecast
            if (upcoming_date is not None) and (d == upcoming_date):
                cur += upcoming_stock
            cur = max(cur, 0.0)
            stocks_current.append(cur)
            if (oos_date is None) and (cur == 0.0):
                oos_date = d

        rop_date = (oos_date - timedelta(days=leadtime + 1)) if oos_date else None
        order_qty = int(math.ceil(forecast * (leadtime + doc))) if forecast > 0 else 0

        base_info = dict(
            CAT=cat, SKU_code=sku,
            **{
                "Available stock": available,
                "Upcoming stock": upcoming_stock,
                "Upcoming date": upcoming_date,
                "Forecast OB/day": forecast,
                "Leadtime (day)": leadtime,
                "DOC": doc,
                "ROP date": rop_date.strftime("%Y-%m-%d") if rop_date else None,
                "Order Qty": order_qty,
            }
        )

        cur_row = {**base_info, **dict(zip(date_cols, stocks_current))}
        out_current.append(cur_row)

        # --- Simulation: Output.ordered ---
        cur2, stocks_ordered = available, []
        arrival_date = (rop_date + timedelta(days=leadtime)) if rop_date else None

        for d in dates:
            if (upcoming_date is not None) and (d == upcoming_date):
                cur2 += upcoming_stock
            if (arrival_date is not None) and (d == arrival_date):
                cur2 += order_qty
            cur2 -= forecast
            cur2 = max(cur2, 0.0)
            stocks_ordered.append(cur2)

        ordered_row = {**base_info, **dict(zip(date_cols, stocks_ordered))}
        out_ordered.append(ordered_row)

    # --- Final Data ---
    input_cols = df.columns.tolist()
    final_cols = input_cols + ["ROP date", "Order Qty"] + date_cols

    return (
        pd.DataFrame(out_current).reindex(columns=final_cols),
        pd.DataFrame(out_ordered).reindex(columns=final_cols),
        date_cols,
        today,
    )


def to_excel_bytes(df_input, out_current, out_ordered):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_input.to_excel(writer, sheet_name="Input", index=False)
        out_current.to_excel(writer, sheet_name="Output.current", index=False)
        out_ordered.to_excel(writer, sheet_name="Output.ordered", index=False)
        
        workbook = writer.book
        pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

        # Xác định SKU có ROP date ở tuần sớm nhất
        df_tmp = out_current.copy()
        df_tmp["ROP_date"] = pd.to_datetime(df_tmp["ROP date"], errors="coerce")
        df_tmp["Week_num"] = df_tmp["ROP_date"].dt.to_period("W-MON")
        earliest_week = df_tmp["Week_num"].min()
        highlight_skus = set(df_tmp.loc[df_tmp["Week_num"] == earliest_week, "SKU_code"])

        def highlight_sheet(ws, df):
            col_sku = list(df.columns).index("SKU_code") + 1
            for row in range(2, ws.max_row + 1):  # bỏ header
                sku = ws.cell(row=row, column=col_sku).value
                if sku in highlight_skus:
                    ws.cell(row=row, column=col_sku).fill = pink_fill

        highlight_sheet(writer.sheets["Output.current"], out_current)
        highlight_sheet(writer.sheets["Output.ordered"], out_ordered)

    buffer.seek(0)
    return buffer

# =============================
# =============================
st.markdown("**Upload** an Excel file (sheet **'Input'**) with columns: `CAT, SKU_code, Available stock, Upcoming stock, Upcoming date, Forecast OB/day, Leadtime (day), DOC`.")

uploaded = st.file_uploader("Upload Input Excel (.xlsx)", type=["xlsx"])
days_ahead = 30

if uploaded:
    try:
        df_in = pd.read_excel(uploaded, sheet_name="Input")
    except Exception as e:
        st.error(f"Error reading file or sheet 'Input': {e}")
        st.stop()

    # Validation
    missing = validate_input(df_in)
    if missing:
        st.error("Missing required columns: " + ", ".join(missing))
        st.stop()

    st.success("✅ Input loaded. Preview:")
    st.dataframe(df_in.head(10))

    if st.button("🚀 Run & Generate Outputs"):
        with st.spinner("Calculating..."):
            out_current, out_ordered, date_cols, used_today = process_input(df_in, days_ahead)

            df_current_display = out_current.copy()
            df_current_display["ROP_date"] = pd.to_datetime(df_current_display["ROP date"], errors="coerce")
            df_current_display["Week_num"] = df_current_display["ROP_date"].dt.to_period("W-MON")

            earliest_week = df_current_display["Week_num"].min()
            highlight_mask = df_current_display["Week_num"] == earliest_week

            def highlight_sku(s):
                return ["background-color: Pink; font-weight: bold;" if highlight_mask.iloc[i] else "" for i in range(len(s))]

            def format_numbers(val):
                if isinstance(val, float) and val.is_integer():
                    return f"{int(val)}"
                elif isinstance(val, float):
                    return f"{val:.2f}"
                return val

            st.subheader("Output.current (preview)")
            st.dataframe(
                df_current_display.style.apply(highlight_sku, subset=["SKU_code"]).format(format_numbers),
                use_container_width=True,
            )

            st.subheader("Output.ordered (preview)")
            st.dataframe(
                df_current_display.style.apply(highlight_sku, subset=["SKU_code"]).format(format_numbers),
                use_container_width=True,
            )

        excel_bytes = to_excel_bytes(df_in, out_current, out_ordered)
        st.download_button(
            "⬇️ Download Results",
            data=excel_bytes,
            file_name=f"Replenishment_Result_{used_today}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.success("✨ Done — Results ready to download!")
else:
    st.info("Please upload an Excel (.xlsx) file containing sheet 'Input'.")
